"""
export/excel_exporter.py

Builds the Excel (.xlsx) report: one sheet with a row per detected date
(file name, page, detected date, normalized date, in-range status,
confidence score), and a summary sheet with the totals required by the
spec. Uses openpyxl directly (rather than pandas.to_excel) so we have
full control over styling — color-coded status cells make the report
much faster to scan in a real business workflow.
"""

from __future__ import annotations

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from utils.logger import get_logger
from utils.models import FileResult, ScanSummary, ValidationStatus

logger = get_logger(__name__)

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

_STATUS_FILLS = {
    ValidationStatus.VALID: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    ValidationStatus.OUT_OF_RANGE: PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    ValidationStatus.NEEDS_REVIEW: PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    ValidationStatus.UNPARSEABLE: PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}

_DETAIL_HEADERS = [
    "File Name", "Page Number", "Detected Date (Raw)", "Normalized Date",
    "Status", "Confidence (%)", "Nearby Label", "Format Inferred", "OCR Engine",
]


def _autosize_columns(ws: Worksheet, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _write_detail_sheet(ws: Worksheet, file_results: list[FileResult]) -> None:
    ws.title = "Date Details"
    ws.append(_DETAIL_HEADERS)
    for col in range(1, len(_DETAIL_HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    for file_result in file_results:
        if file_result.error:
            ws.append([file_result.file_name, "—", "—", "—", f"ERROR: {file_result.error}", "—", "—", "—", "—"])
            for col in range(1, len(_DETAIL_HEADERS) + 1):
                ws.cell(row=row_idx, column=col).fill = _STATUS_FILLS[ValidationStatus.UNPARSEABLE]
            row_idx += 1
            continue

        for page in file_result.pages:
            if page.error:
                ws.append([file_result.file_name, page.page_number, "—", "—", f"ERROR: {page.error}", "—", "—", "—", "—"])
                for col in range(1, len(_DETAIL_HEADERS) + 1):
                    ws.cell(row=row_idx, column=col).fill = _STATUS_FILLS[ValidationStatus.UNPARSEABLE]
                row_idx += 1
                continue

            for detected in page.dates:
                ws.append([
                    file_result.file_name,
                    page.page_number,
                    detected.raw_text,
                    detected.display_date,
                    detected.status.value,
                    detected.confidence,
                    detected.nearby_label or "—",
                    "Yes" if detected.format_inferred else "No",
                    detected.engine.value,
                ])
                fill = _STATUS_FILLS.get(detected.status)
                if fill:
                    for col in range(1, len(_DETAIL_HEADERS) + 1):
                        ws.cell(row=row_idx, column=col).fill = fill
                row_idx += 1

    _autosize_columns(ws, [28, 12, 20, 16, 20, 14, 20, 14, 12])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_DETAIL_HEADERS))}{row_idx - 1}"

    confidence_col = _DETAIL_HEADERS.index("Confidence (%)") + 1
    for row in range(2, row_idx):
        cell = ws.cell(row=row, column=confidence_col)
        if isinstance(cell.value, (int, float)):
            cell.number_format = "0.0"


def _write_summary_sheet(ws: Worksheet, summary: ScanSummary, start_date_str: str, end_date_str: str) -> None:
    ws.title = "Summary"
    rows = [
        ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Valid Date Range", f"{start_date_str} to {end_date_str}"),
        ("", ""),
        ("Total Files Scanned", summary.total_files),
        ("Total Pages Processed", summary.total_pages),
        ("Total Dates Detected", summary.total_dates_detected),
        ("Dates Outside Range", summary.total_out_of_range),
        ("Dates Needing Manual Review", summary.total_needs_review),
        ("Average OCR Confidence (%)", summary.average_confidence),
        ("Total Processing Time (seconds)", summary.processing_seconds),
    ]
    for label, value in rows:
        ws.append([label, value])

    for row in range(1, len(rows) + 1):
        ws.cell(row=row, column=1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 30


def export_report(file_results: list[FileResult], summary: ScanSummary,
                   start_date_str: str, end_date_str: str, output_path: str) -> str:
    """Build and save the full Excel report. Returns the output path."""
    wb = Workbook()
    summary_ws = wb.active
    _write_summary_sheet(summary_ws, summary, start_date_str, end_date_str)

    detail_ws = wb.create_sheet("Date Details")
    _write_detail_sheet(detail_ws, file_results)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    logger.info("Excel report saved to %s", output_path)
    return output_path
